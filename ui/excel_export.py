"""coordinates.xlsx writer -- a convenience export, NOT a contract output.

D35 (plan.md): C6 is frozen at exactly 16 GeoJSON properties and this
spreadsheet must not present itself as an alternative authority. Everything
here is DERIVED from files the run already wrote -- the GeoJSON (geometry,
lat/lon, scores, confidence), the mask TIFF (per-pixel sheet, its CRS and
transform), and run_manifest.json (the Metadata sheet, including git_sha so
any exported file traces back to the exact run that produced it).

No tkinter import at module level: this must stay testable headless
(docs/ui_plan.md section 5).
"""
from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import rasterio
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rasterio.warp import transform as warp_transform

# A pathological threshold can mark hundreds of thousands of pixels; the
# per-pixel sheet is the one place a runaway row count would freeze both the
# UI worker and Excel itself. Above this cap the sheet records that it was
# truncated rather than silently writing a subset.
MAX_PIXEL_ROWS = 200_000


def pixels_to_latlon(transform, crs, rows, cols) -> tuple[np.ndarray, np.ndarray]:
    """Pixel indices -> WGS84 (lat, lon). The ONLY reprojection in the UI,
    mirroring geospatial.projections' EPSG:4326 target."""
    rows = np.asarray(rows, dtype=np.int64)
    cols = np.asarray(cols, dtype=np.int64)
    xs, ys = rasterio.transform.xy(transform, rows, cols, offset="center")
    src = rasterio.crs.CRS.from_user_input(crs)
    dst = rasterio.crs.CRS.from_epsg(4326)
    if src == dst:
        lons, lats = np.asarray(xs, dtype=float), np.asarray(ys, dtype=float)
    else:
        lons, lats = warp_transform(src, dst,
                                    [float(x) for x in xs],
                                    [float(y) for y in ys])
        lons, lats = np.asarray(lons), np.asarray(lats)
    return lats, lons


_REGION_COLUMNS = [
    "ROI ID", "Center Latitude", "Center Longitude",
    "Area (m2)", "Perimeter (m)",
    "Anomaly Score", "Confidence", "Confidence Components",
    "Timestamp", "Source Scene", "Class",
]

_PIXEL_COLUMNS = ["Pixel Row", "Pixel Column", "Latitude", "Longitude"]

_GEO_FIELDS = ("roi_id", "lat", "lon", "area", "perimeter", "anomaly_score",
               "confidence", "confidence_components", "timestamp",
               "source_scene", "class")


def _metadata_rows(manifest: dict) -> list[list]:
    return [
        ["Field", "Value"],
        ["Note", "Convenience export derived from the run's GeoJSON/mask/manifest; "
                 "the GeoJSON remains the authoritative output (D35)"],
        ["Scene", manifest.get("scene")],
        ["Source", manifest.get("source")],
        ["Detector", manifest.get("detector")],
        ["Threshold percentile", manifest.get("threshold_pct")],
        ["Profile", manifest.get("profile")],
        ["Normalize method", manifest.get("normalize_method")],
        ["Detector params", json.dumps(manifest.get("detector_params", {}))],
        ["Window (row_off, col_off, h, w)", str(manifest.get("window"))],
        ["ROIs detected", manifest.get("n_rois")],
        ["Git SHA", manifest.get("git_sha")],
        ["Package versions", json.dumps(manifest.get("package_versions", {}))],
        ["Stage timings (s)", json.dumps(manifest.get("timings_s", {}))],
    ]


def export_coordinates_xlsx(xlsx_path: str | Path, geojson_path: str | Path,
                            manifest_path: str | Path) -> Path:
    """Write coordinates.xlsx from a finished run's outputs. Returns xlsx path."""
    xlsx_path = Path(xlsx_path)
    manifest = json.loads(Path(manifest_path).read_text())
    features = json.loads(Path(geojson_path).read_text()).get("features", [])

    wb = Workbook()

    # --- Sheet 1: regions, straight from the GeoJSON features -------------
    ws_regions = wb.active
    ws_regions.title = "Anomaly Regions"
    ws_regions.append(_REGION_COLUMNS)
    for feat in features:
        p = feat.get("properties", feat)   # gpd writes nested properties
        ws_regions.append([
            p.get("roi_id"), p.get("lat"), p.get("lon"),
            p.get("area"), p.get("perimeter"),
            p.get("anomaly_score"), p.get("confidence"),
            ",".join(p.get("confidence_components") or []),
            p.get("timestamp"), p.get("source_scene"), p.get("class"),
        ])

    # --- Sheet 2: per-pixel rows, from the run's mask TIFF -----------------
    ws_pixels = wb.create_sheet("Anomaly Pixels")
    ws_pixels.append(_PIXEL_COLUMNS + ["Score (normalized)"])
    mask_tif = (manifest.get("outputs") or {}).get("mask")
    truncated = False
    if mask_tif and Path(mask_tif).exists():
        with rasterio.open(mask_tif) as ds:
            mask = ds.read(1)
            transform, crs = ds.transform, ds.crs
        rows, cols = np.where(mask == 1)
        if rows.size > MAX_PIXEL_ROWS:
            rows, cols = rows[:MAX_PIXEL_ROWS], cols[:MAX_PIXEL_ROWS]
            truncated = True
        norm_tif = (manifest.get("outputs") or {}).get("anom_norm")
        scores = None
        if norm_tif and Path(norm_tif).exists():
            with rasterio.open(norm_tif) as ds:
                scores = ds.read(1)
        if rows.size:
            lats, lons = pixels_to_latlon(transform, crs, rows, cols)
            for i, (r, c) in enumerate(zip(rows, cols)):
                score = float(scores[r, c]) if scores is not None else None
                ws_pixels.append([int(r), int(c), float(lats[i]), float(lons[i]), score])

    # --- Sheet 3: metadata from the run manifest ---------------------------
    ws_meta = wb.create_sheet("Metadata")
    for row in _metadata_rows(manifest):
        ws_meta.append(row)
    if truncated:
        ws_meta.append(["Pixel sheet truncated at", MAX_PIXEL_ROWS])

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for ws in (ws_regions, ws_pixels, ws_meta):
        ws.freeze_panes = "A2"
        if ws.max_row >= 1:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for column_cells in ws.columns:
            width = max((len(str(c.value)) for c in column_cells if c.value is not None),
                        default=8)
            ws.column_dimensions[column_cells[0].column_letter].width = min(width + 2, 40)

    wb.save(xlsx_path)
    return xlsx_path
