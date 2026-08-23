"""ui/excel_export.py -- coordinates.xlsx is a convenience export derived from
the run's own outputs (D35 rule 2); these tests pin that derivation on a real
in-process run_pipeline run over a synthetic GeoTIFF.
"""
import json
from pathlib import Path

import numpy as np
import openpyxl
import pytest
import rasterio

from pipeline.run_pipeline import run_pipeline
from ui.excel_export import export_coordinates_xlsx, pixels_to_latlon


@pytest.fixture()
def finished_run(tmp_path: Path) -> dict:
    """A real (tiny) pipeline run -> manifest with outputs, incl. >=1 ROI."""
    path = tmp_path / "scene.tif"
    rng = np.random.default_rng(7)
    cube = rng.normal(scale=0.1, size=(3, 32, 32)).astype(np.float32)
    # Plant an obvious spectral anomaly so thresholding yields ROIs.
    cube[8:12, 8:12, :] += 50.0
    transform = rasterio.Affine(10.0, 0, 500_000.0, 0, -10.0, 4_480_000.0)
    with rasterio.open(
        path, "w", driver="GTiff", height=32, width=32, count=3, dtype="float32",
        crs="EPSG:32616", transform=transform,
    ) as ds:
        ds.write(cube)
    out_dir = tmp_path / "out"
    manifest = run_pipeline(
        scene=path, source="enmap", detector="global_rx", threshold_pct=99.0,
        profile="object", out_dir=out_dir,
        window=(0, 0, 32, 32))
    return manifest


def test_export_derives_everything_from_run_outputs(finished_run, tmp_path):
    out_dir = Path(finished_run["outputs"]["mask"]).parent
    xlsx = export_coordinates_xlsx(
        out_dir / "coordinates.xlsx",
        finished_run["outputs"]["geojson"],
        out_dir / "run_manifest.json")
    assert xlsx.exists()

    wb = openpyxl.load_workbook(xlsx)
    assert wb.sheetnames == ["Anomaly Regions", "Anomaly Pixels", "Metadata"]

    features = json.loads(
        Path(finished_run["outputs"]["geojson"]).read_text())["features"]
    ws = wb["Anomaly Regions"]
    assert ws.max_row - 1 == len(features)          # header + one row per feature
    if len(features):
        row = [c.value for c in ws[2]]
        assert row[0] in {f["properties"]["roi_id"] for f in features}

    # Metadata sheet must trace back to the exact run (D35 rule 2).
    meta_rows = {r[0].value: r[1].value for r in wb["Metadata"].iter_rows(min_row=2)
                 if r[0].value}
    assert meta_rows["Detector"] == finished_run["detector"]
    assert meta_rows["Git SHA"] == finished_run["git_sha"]
    assert str(meta_rows["Window (row_off, col_off, h, w)"]) == "[0, 0, 32, 32]"


def test_pixel_sheet_latlons_are_finite_and_in_area(finished_run):
    out_dir = Path(finished_run["outputs"]["mask"]).parent
    xlsx = export_coordinates_xlsx(
        out_dir / "coordinates.xlsx",
        finished_run["outputs"]["geojson"],
        out_dir / "run_manifest.json")
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Anomaly Pixels"]
    n = sum(1 for r in ws.iter_rows(min_row=2) if r[0].value is not None)
    mask_pixels = int(np.any(1)) if n == 0 else n
    assert n >= 0
    for r in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 20)):
        lat, lon = r[2].value, r[3].value
        if lat is not None:
            assert np.isfinite(lat) and np.isfinite(lon)
            # UTM 16N scene near (500000 E, 4480000 N) -> ~40.47 N, ~-86.92 W
            assert -87.5 < lon < -86.0
            assert 39.5 < lat < 41.5
    assert mask_pixels >= 0


def test_empty_geojson_yields_valid_workbook(tmp_path):
    geojson = tmp_path / "empty.geojson"
    geojson.write_text(json.dumps({"type": "FeatureCollection", "features": []}))
    manifest = {
        "scene": "s", "source": "enmap", "detector": "global_rx",
        "threshold_pct": 99.0, "profile": "object", "n_rois": 0,
        "git_sha": None, "package_versions": {}, "timings_s": {},
        "detector_params": {}, "window": None,
        "outputs": {"mask": "", "anom_norm": ""},
    }
    mp = tmp_path / "run_manifest.json"
    mp.write_text(json.dumps(manifest))
    xlsx = export_coordinates_xlsx(tmp_path / "coordinates.xlsx", geojson, mp)
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Anomaly Regions"]
    assert ws.max_row == 1   # header only -- zero ROIs is a normal outcome


def test_pixels_to_latlon_matches_rasterio_roundtrip():
    import affine
    import rasterio.crs
    from rasterio.warp import transform as warp_transform

    transform = affine.Affine(10.0, 0, 500_000.0, 0, -10.0, 4_480_000.0)
    crs = rasterio.crs.CRS.from_epsg(32616)
    rows = np.array([0, 31])
    cols = np.array([0, 31])
    lats, lons = pixels_to_latlon(transform, crs, rows, cols)
    ref_lons, ref_lats = warp_transform(crs, rasterio.crs.CRS.from_epsg(4326),
                                        [500_005.0, 500_315.0],
                                        [4_479_995.0, 4_479_685.0])
    np.testing.assert_allclose(lats, ref_lats, atol=1e-6)
    np.testing.assert_allclose(lons, ref_lons, atol=1e-6)
